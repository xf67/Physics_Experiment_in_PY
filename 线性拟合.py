import os
import sys
import platform
try:
    import numpy as np
    import matplotlib.pyplot as plt
except:
    print("缺少numpy和matplotlib库.")
    print("尝试pip3 install numpy以及pip3 install matplotlib")
    sys.exit()
try:
    import openpyxl as xl
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Font 
    no_excel=0
except:
    print("缺少openpyxl库,将无法实现和excel的连接")
    print("尝试pip3 install openpyxl")
    no_excel=1

x=[]
y=[]

class Excelrw:
    def __init__(self,fpath,new_excel=0):
        self.fpath =fpath
        self.new_excel=new_excel
        self.cla_x =""
        self.cla_y =""
        self.cla_sheet =""
        self.open_ok =0
        self.cla_head =""
        self.cla_foot =""
    def open_book(self):
        while(not self.open_ok):
            self.fpath=input("输入文件地址(示例:/Users/admin/Desktop/test.xlsx)\n注:可以直接把文件拖到命令行窗口\n")
            self.fpath=self.fpath.replace("'",'')
            if not os.path.exists(self.fpath):
                print("文件地址错误")
            else:
                if not self.fpath.endswith(".xlsx" or ".XLSX"):
                    print("不是XLSX文件.\t注:如果是XLS请先转换为XLSX.")
                elif self.fpath.endswith(".xlsx" or ".XLSX"):
                    self._openbook()
                    self.open_ok=1
    def _openbook(self):
        print(f'加载 {self.fpath}')
        try:
            self.wb1 = load_workbook(self.fpath)
        except:
            print("文件读取出错")
            sys.exit()
        if self.new_excel==0:
            self._set()
            self._read()
    def _set(self):
        self.sheets = self.wb1.sheetnames
        print(self.sheets,end="")
        tmp_flag=1
        while(self.cla_sheet=="" and tmp_flag==1):
            self.cla_sheet = input("这是此文件中所含的数据表,\n请选择要分析的表格,用1,2,3...表示,或回车以选择默认值1: ")
            if self.cla_sheet=="":
                self.cla_sheet="1"
                tmp_flag=0
        self.wb=self.wb1[self.sheets[int(self.cla_sheet)-1]]
        while(self.cla_x==""):
            self.cla_x = input("选择x所在的列: ")
        while(self.cla_y==""):
            self.cla_y = input("选择y所在的列: ")
        while(self.cla_head==""):
            self.cla_head = input("选择具体数据开始的行号: ")
        tmp_foot = self.wb.max_row
        tmp_foot_ok=""
        while(tmp_foot_ok==""):
            tmp_foot_ok = input(f"自动检测数据的最后一行为{tmp_foot},输入Y表示认可,或输入行号来修改: ")
            if tmp_foot_ok=="Y" or tmp_foot_ok=="y":
                self.cla_foot = tmp_foot
            else :
                self.cla_foot = tmp_foot_ok
    def _read(self):
        for row in range(int(self.cla_head),int(self.cla_foot)+1):
            y.append(self.wb[self.cla_y+str(row)].value)
            x.append(self.wb[self.cla_x+str(row)].value)
    def save(self):
        if 'Result' in self.wb1.sheetnames:
            del self.wb1['Result']
        self.wb1.create_sheet('Result')
        self.wbok=self.wb1['Result']
        self.wbok['B1']='最终数据'
        self.wbok['B2']=last
        self.wbok['C3']='x'
        self.wbok['D3']='y'
        for i in range(0,len(x)):
            self.wbok['C'+str(i+4)]=x[i]
            self.wbok['D'+str(i+4)]=y[i]
        self.wbok['E1']='原数据'
        self.wbok['E2']=raw
        self.wbok['F3']='x'
        self.wbok['G3']='y'
        for i in range(0,len(x)):
            self.wbok['F'+str(i+4)]=rawx[i]
            self.wbok['G'+str(i+4)]=rawy[i]
        try:
            self.wb1.save(self.fpath)
        except PermissionError:
            print("文件被占用,把excel关掉试试.")

def inputx():
    tmp=0
    while(1):
        iin=input(f"第{tmp+1}行\tx=")
        if iin=='e' or iin=='E':
            break
        try:
            fin=float(iin)
            x.append(fin)
            tmp+=1
        except:
            print("输入非法")
def inputy():
    for i in range(0,len(x)):
        iin=input(f"第{i+1}行\tx={x[i]}\ty=")
        try:
            fin=float(iin)
            y.append(fin)
        except:
            print("输入非法")
def clearall():
    sys=platform.system()
    if sys=='Darwin' or sys=="Linux":
        os.system('clear')
    elif sys=='Windows':
        os.system('cls')
def printall(xx,yy):
    print(" \tx\ty")
    for i in range(0,len(xx)):
        print(f"{i+1}\t{xx[i]}\t{yy[i]}")
def modify():
    modif=input("需要修改? Y/N :")
    if modif=='N' or modif=="n":
        return 
    elif modif=='Y' or modif=='y':
        while(1):
            col=input("改x还是y? x/y :")
            if col=='x':
                a=x
                break
            elif col=='y':
                a=y
                break
            else:
                print("输入非法")
        while(1):   
            row=input("第几行:")
            try:
                row=int(row)
            except:
                print("输入非法")
            if row in range(1,len(x)+1):
                break
            else:
                print("输入非法")
        while(1):
            what=input(f"将{a[row-1]}改为:")
            try:
                what=float(what)
                a[row-1]=what
                clearall()
                printall(x,y)
                modify()
                break
            except:
                print("输入非法")
def draw():
    X=np.arange(min(x),max(x),(max(x)-min(x))/200)
    Y=k*X+b
    plt.plot(X, Y, color='red')
    plt.scatter(x, y, color='black',marker='s')
    plt.show()
def calculate(x1,y1):
    global k
    global b
    global pearsonr
    global delta_b
    global delta_k
    global delta_y
    global rr
    global serror
    axy,ad,_,_,_=np.polyfit(x1,y1,1,full=True)
    serror=ad[0]
    k=axy[0]
    b=axy[1]
    y_mean=sum(y1)/len(y1)
    s_list=[pow(ys-y_mean,2) for ys in y1 ]
    stotal=sum(s_list)
    rr=1-serror/stotal
    pearsonr=pow(rr,1/2)
    x2_list=[pow(xs,2) for xs in x1 ]
    x2_sum=sum(x2_list)
    x2_mean=x2_sum/len(x)
    x_mean=sum(x)/len(x)
    delta_y=pow(serror/(len(x1)-1),1/2)
    delta_k=delta_y/pow(x2_mean-pow(x_mean,2),1/2)
    delta_b=pow(x2_sum,1/2)*delta_k
def evil(xt,yt):
    xk=xt[:]
    yk=yt[:]
    evi=[]
    ok=0.999
    calculate(xk,yk)
    if len(xt)<=3:
        return
    ttmp=input("Run the evil check? Y/N: ")
    if ttmp=='n' or ttmp=='N':
        return
    elif ttmp=='y' or ttmp=='Y':
        if rr>0.999:
            while(1):
                tmp2=input("R值已经足够,给出期望的R值的9的个数:")
                try:
                    tmp2=int(tmp2)
                    ok=1-pow(0.1,tmp2)
                    break
                except:
                    print("输入非法")
        times=0
        while(rr<ok and times<3):
            print(f"第{times+1}次迭代...")
            times+=1
            arrY=[xx*k+b for xx in xk]
            dy=[abs(yk[i]-arrY[i]) for i in range(0,len(yk))]
            bad_index=dy.index(max(dy))
            new_x=xk[:]
            new_y=yk[:]
            del new_x[bad_index]
            del new_y[bad_index]
            calculate(new_x,new_y)
            good_y=xk[bad_index]*k+b
            new_y=yk[:]
            new_x=xk[:]
            new_y[bad_index]=good_y
            calculate(xk,new_y)
            evi.append(f"建议:第{bad_index+1}行\tx={xk[bad_index]}\ty={yk[bad_index]}->y={good_y}")
            yk[bad_index]=good_y
        clearall()
        printall(x,y)
        #print(raw)
        print("____________________")
        if not evi:
            print("无建议")
        for evis in evi:
            print(evis)
        print(f"斜率={k}+-{delta_k}\n截距={b}+-{delta_b}\n残差平方和={serror}\nPearson's r={pearsonr}\nR平方(COD)={rr}")
        modify()
        evil(x,y)
        return
        
def main():
    global raw
    global rawx
    global rawy
    global last
    print("\t\n\
          _     _                         _____ _ _\n \
        | |   (_)_ __   ___  __ _ _ __  |  ___(_) |_\n \
        | |   | | '_ \ / _ \/ _` | '__| | |_  | | __|\n \
        | |___| | | | |  __/ (_| | |    |  _| | | |_ \n \
        |_____|_|_| |_|\___|\__,_|_|    |_|   |_|\__|\n\
        \t\t powered by 67")
    input("回车以继续")
    clearall()
    while(1):
        exc=input("输入模式:1=手动输入,2=excel导入 :")
        if exc=='1':
            print("输入x,以e结束")
            inputx()
            clearall()
            print("输入y")
            inputy()
            break
        elif exc=='2' and no_excel==0:
            excelrw=Excelrw("",0)
            excelrw.open_book()
            break
        elif exc=='2' and no_excel==1:
            print("缺少openpyxl库,只能手动输入")
            print("输入x,以e结束")
            inputx()
            clearall()
            print("输入y")
            inputy()
            break
        else:
            print("输入非法")
    clearall()
    printall(x,y)
    modify()
    clearall()
    printall(x,y)
    calculate(x,y)
    raw=(f"斜率={k}+-{delta_k}\n截距={b}+-{delta_b}\n残差平方和={serror}\nPearson's r={pearsonr}\nR平方(COD)={rr}")
    rawx=x[:]
    rawy=y[:]
    print(raw)
    draw()
    evil(x,y)
    clearall()
    if set(y).difference(set(rawy)):
        print("原数据:")
        printall(rawx,rawy)
        print(raw)
        print("\n修改后:")
    printall(x,y)
    calculate(x,y)
    last=(f"斜率={k}+-{delta_k}\n截距={b}+-{delta_b}\n残差平方和={serror}\nPearson's r={pearsonr}\nR平方(COD)={rr}")
    print(last)
    if exc=='2':
        excelrw.save()
        print("数据已经保存在同一文件中名为Result的数据表中")
    elif exc=='1' and no_excel==0:
        need_excel=input("需要保存到excel吗? Y/N: ")
        if need_excel=='y' or need_excel=='Y':
            excelrw=Excelrw("",1)
            excelrw.open_book()
            excelrw.save()
            print("数据已经保存在名为Result的数据表中")
    print("\n运行结束,以上为最终结果.")
    draw()
    
if __name__ == '__main__':
    main()


