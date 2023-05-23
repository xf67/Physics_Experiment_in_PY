#-*- coding: utf-8 -*-
import numpy as np
import math
import matplotlib.pyplot as plt
import openpyxl as xl
from openpyxl import load_workbook

u=[] ; l=[]

class Excelrw:
    def __init__(self,fpath):
        self.fpath =fpath
        self.cla_x =""
        self.cla_y =""
        self.cla_sheet =""
        self.open_ok =0
        self.cla_head =""
        self.cla_foot =""
    def open_book(self):
        while(not self.open_ok):
            self.fpath=input("excel文件地址: ")
            self.fpath=self.fpath.replace("'",'')
            self._openbook()
            self.open_ok=1
    def _openbook(self):
        self.wb1 = load_workbook(self.fpath)
        self._set()
        self._read()
    def _set(self):
        self.sheets = self.wb1.sheetnames
        self.wb=self.wb1[self.sheets[0]]
        self.cla_x = input("l所在的列: ")
        self.cla_y = input("U所在的列: ")
        self.cla_head = input("数据开始的行号: ")
        self.cla_foot = self.wb.max_row
    def _read(self):
        for row in range(int(self.cla_head),int(self.cla_foot)+1):
            l.append(float(self.wb[self.cla_x+str(row)].value))
            u.append(float(self.wb[self.cla_y+str(row)].value))
            
excelrw=Excelrw("")
excelrw.open_book()
x=[]
b_expri=[]
K=float(input("灵敏度K:"))

for ll in l:
    x.append((ll-15.70)/100)
for uu in u:
    b_expri.append(uu/1000/K)

def b_theo(xx):
    NU=4*math.pi*1e-7
    N=3000
    L=0.26
    IM=0.25
    D=range(25+1,45+1,2)
    b=0
    for dd in D:
        dd=dd/1000
        b=b+NU*N/L*IM*((L+2*xx)/2/((dd**2+(L+2*xx)**2)**(1/2))+(L-2*xx)/2/((dd**2+(L-2*xx)**2)**(1/2)))
    return b/10

def draw():
    plt.rcParams['font.sans-serif'] = 'MS Mincho'
    #因为没有安装字体,我的电脑在绘图时不能正确显示简体中文宋体字,所以这里用了日本汉字明体字
    X=np.arange(-0.18,0.16,0.32/200)
    Y=[]
    for xx in X:
        Y.append(b_theo(xx))
    plt.plot(X, Y, color='red',label="理論値")
    plt.scatter(x, b_expri, color='black',marker='s',s = 14)
    plt.plot(x, b_expri, color='black',label="実験値")
    plt.xlabel("x(m)")
    plt.ylabel("B(T)")
    plt.title("通電螺線管的磁場分布")
    plt.legend(loc='best')
    print("图象生成完毕")
    
    plt.show()

draw()


